from __future__ import annotations

from .domain import Assignment, MonthInput
from .domain import SLOT_LABEL_JA, SLOT_ORDER


def export_xlsx(mi: MonthInput, assignments: tuple[Assignment, ...], out_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ModuleNotFoundError as e:
        raise RuntimeError(
            "openpyxl が見つかりません。`pip install -r requirements.txt` を実行してください。"
        ) from e

    staff_by_id = mi.staff_by_id()
    wb = Workbook()
    ws = wb.active
    ws.title = mi.month

    header = ["日付", "曜日", "種別"] + [SLOT_LABEL_JA[s] for s in SLOT_ORDER] + ["マネージャー有"]
    ws.append(header)

    fill_header = PatternFill("solid", fgColor="1F2937")  # dark gray
    font_header = Font(color="FFFFFF", bold=True)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    weekdays = "月火水木金土日"
    for a in assignments:
        d = a.day
        has_mgr = any(staff_by_id[sid].is_manager for sid in a.slots.values())
        kind = "土曜" if d.weekday() == 5 else "平日"

        row = [d.isoformat(), weekdays[d.weekday()], kind]
        for slot_name in SLOT_ORDER:
            sid = a.slots.get(slot_name)
            row.append(staff_by_id[sid].name if sid else "")
        row.append("OK" if has_mgr else "NG")
        ws.append(row)

    # Simple sizing
    ws.freeze_panes = "A2"
    widths = [12, 6, 8] + [12] * len(SLOT_ORDER) + [16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w

    # Center align body
    for r in range(2, len(assignments) + 2):
        for c in range(1, len(header) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    wb.save(out_path)
