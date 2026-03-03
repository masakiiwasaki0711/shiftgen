from __future__ import annotations

from .calendar_utils import is_saturday
from .domain import Assignment, MonthInput, SLOT_LABEL_JA, SLOT_ORDER

HOURS_WEEKDAY = 8.5
HOURS_SATURDAY = 4.5


def compute_hours(
    mi: MonthInput, assignments: tuple[Assignment, ...]
) -> list[tuple[str, str, bool, int, int, float]]:
    """スタッフごとの勤務回数・合計時間を返す。

    Returns:
        list of (staff_id, name, is_manager, wd_count, sat_count, total_hours)
        スタッフ登録順に並ぶ。
    """
    wd_count: dict[str, int] = {s.id: 0 for s in mi.staff}
    sat_count: dict[str, int] = {s.id: 0 for s in mi.staff}

    for a in assignments:
        sat = is_saturday(a.day)
        for sid in a.slots.values():
            if sat:
                sat_count[sid] += 1
            else:
                wd_count[sid] += 1

    result = []
    for s in mi.staff:
        wd = wd_count[s.id]
        sat = sat_count[s.id]
        total = wd * HOURS_WEEKDAY + sat * HOURS_SATURDAY
        result.append((s.id, s.name, s.is_manager, wd, sat, total))
    return result


def export_xlsx(mi: MonthInput, assignments: tuple[Assignment, ...], out_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ModuleNotFoundError as e:
        raise RuntimeError(
            "openpyxl が見つかりません。`pip install -r requirements.txt` を実行してください。"
        ) from e

    staff_by_id = mi.staff_by_id()
    wb = Workbook()
    ws = wb.active
    ws.title = mi.month

    header = ["日付", "曜日", "種別"] + [SLOT_LABEL_JA[s] for s in SLOT_ORDER] + ["マネージャー有"]
    ws.append(header)

    fill_header = PatternFill("solid", fgColor="1F2937")
    font_header = Font(color="FFFFFF", bold=True)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    weekdays = "月火水木金土日"
    for a in assignments:
        d = a.day
        has_mgr = any(staff_by_id[sid].is_manager for sid in a.slots.values())
        kind = "土曜" if d.weekday() == 5 else "平日"

        row = [d.isoformat(), weekdays[d.weekday()], kind]
        for slot_name in SLOT_ORDER:
            sid = a.slots.get(slot_name)
            row.append(staff_by_id[sid].name if sid else "")
        row.append("OK" if has_mgr else "NG")
        ws.append(row)

    ws.freeze_panes = "A2"
    widths = [12, 6, 8] + [12] * len(SLOT_ORDER) + [16]
    from openpyxl.utils import get_column_letter

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in range(2, len(assignments) + 2):
        for c in range(1, len(header) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    # --- 2枚目: 勤務時間集計シート ---
    ws2 = wb.create_sheet(title="勤務時間集計")
    summary_header = ["名前", "マネージャー", "平日勤務回数", f"平日時間(×{HOURS_WEEKDAY}h)", "土曜勤務回数", f"土曜時間(×{HOURS_SATURDAY}h)", "合計時間(h)"]
    ws2.append(summary_header)
    for col in range(1, len(summary_header) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    hours_data = compute_hours(mi, assignments)
    for _sid, name, is_mgr, wd, sat, total in hours_data:
        ws2.append([
            name,
            "○" if is_mgr else "",
            wd,
            wd * HOURS_WEEKDAY,
            sat,
            sat * HOURS_SATURDAY,
            total,
        ])

    # 合計行
    n = len(hours_data)
    if n > 0:
        ws2.append([
            "合計",
            "",
            sum(r[3] for r in hours_data),
            sum(r[3] * HOURS_WEEKDAY for r in hours_data),
            sum(r[4] for r in hours_data),
            sum(r[4] * HOURS_SATURDAY for r in hours_data),
            sum(r[5] for r in hours_data),
        ])
        total_row = n + 2
        font_total = Font(bold=True)
        for col in range(1, len(summary_header) + 1):
            ws2.cell(row=total_row, column=col).font = font_total

    summary_widths = [14, 14, 14, 20, 14, 20, 14]
    for i, w in enumerate(summary_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, len(hours_data) + 3):
        for c in range(1, len(summary_header) + 1):
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(out_path)

