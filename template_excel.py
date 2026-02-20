from __future__ import annotations

from dataclasses import dataclass
from datetime import date

from domain import MonthInput, Requirements, Staff
from calendar_utils import month_range


def _parse_date(s: str) -> date:
    y, m, d = s.split("-")
    return date(int(y), int(m), int(d))


def _as_date_cell(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        return _parse_date(v.strip())
    raise ValueError(f"日付の形式が不正です: {v!r}")


def export_template_xlsx(mi: MonthInput, out_path: str) -> None:
    """
    Template format (sheets):
    - Config: month, auto_close_jp_holidays, saturday_max_per_person, prefer_max_headcount
    - Staff: id, name, is_manager, allowed_kinds (comma-separated, optional)
    - RequestsOffCalendar: staff_id, name, day columns (1..31). Mark OFF with "OFF" or "1".
    - RequestsOff (legacy): staff_id, date (YYYY-MM-DD). 1 row per request.
    - Closed: date (YYYY-MM-DD). 1 row per manual close. (Auto JP holidays are handled by flag)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as e:
        raise RuntimeError(
            "openpyxl が見つかりません。`pip install -r requirements.txt` を実行してください。"
        ) from e

    wb = Workbook()
    wb.remove(wb.active)

    fill_header = PatternFill("solid", fgColor="1F2937")
    font_header = Font(color="FFFFFF", bold=True)

    def write_header(ws, cols: list[str]):
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    # Config
    ws = wb.create_sheet("Config")
    write_header(ws, ["key", "value"])
    ws.append(["month", mi.month])
    ws.append(["auto_close_jp_holidays", "TRUE" if mi.auto_close_jp_holidays else "FALSE"])
    ws.append(["saturday_max_per_person", str(mi.requirements.saturday_max_per_person)])
    ws.append(["prefer_max_headcount", "TRUE" if mi.requirements.prefer_max_headcount else "FALSE"])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    # Staff
    ws = wb.create_sheet("Staff")
    write_header(ws, ["id", "name", "is_manager", "allowed_kinds"])
    for s in mi.staff:
        allowed = ",".join(s.allowed_kinds) if s.allowed_kinds else ""
        ws.append([s.id, s.name, "TRUE" if s.is_manager else "FALSE", allowed])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30

    # RequestsOff
    start, end = month_range(mi.month)
    days = list(range(1, end.day + 1))

    ws = wb.create_sheet("RequestsOffCalendar")
    write_header(ws, ["staff_id", "name"] + [str(d) for d in days])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for col in range(3, 3 + len(days)):
        ws.column_dimensions[get_column_letter(col)].width = 4
    # One row per staff, blank by default.
    for s in mi.staff:
        ws.append([s.id, s.name] + [""] * len(days))

    # Legacy sheet kept for compatibility with existing files/tools.
    ws = wb.create_sheet("RequestsOff")
    write_header(ws, ["staff_id", "date"])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14

    # Closed
    ws = wb.create_sheet("Closed")
    write_header(ws, ["date"])
    for d in mi.closed_dates:
        ws.append([d.isoformat()])
    ws.column_dimensions["A"].width = 14

    wb.save(out_path)


def import_from_template_xlsx(path: str) -> MonthInput:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as e:
        raise RuntimeError(
            "openpyxl が見つかりません。`pip install -r requirements.txt` を実行してください。"
        ) from e

    wb = load_workbook(path)
    if "Config" not in wb.sheetnames or "Staff" not in wb.sheetnames:
        raise ValueError("テンプレのシート構成が不正です (Config/Staff が必要)。")

    # Config
    config = {}
    ws = wb["Config"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        key = str(row[0]).strip()
        val = row[1]
        config[key] = val

    month = str(config.get("month", "")).strip()
    if not month or len(month) != 7 or month[4] != "-":
        raise ValueError("Config.month が不正です (YYYY-MM)。")

    def as_bool(v, default: bool) -> bool:
        if v is None or v == "":
            return default
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if s in ("true", "1", "yes", "y"):
            return True
        if s in ("false", "0", "no", "n"):
            return False
        return default

    auto_close = as_bool(config.get("auto_close_jp_holidays"), True)
    prefer_max = as_bool(config.get("prefer_max_headcount"), True)
    sat_max = int(str(config.get("saturday_max_per_person", "3")).strip())
    requirements = Requirements(saturday_max_per_person=sat_max, prefer_max_headcount=prefer_max)

    # Staff
    staff: list[Staff] = []
    ws = wb["Staff"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        sid = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] is not None else ""
        if not name:
            raise ValueError(f"Staff.name が空です: {sid}")
        is_mgr = as_bool(row[2], False)
        allowed_raw = str(row[3]).strip() if (len(row) > 3 and row[3] is not None) else ""
        allowed = tuple(x.strip() for x in allowed_raw.split(",") if x.strip()) or None
        staff.append(Staff(id=sid, name=name, is_manager=is_mgr, allowed_kinds=allowed))
    if not staff:
        raise ValueError("Staff シートにスタッフがありません。")

    staff_ids = {s.id for s in staff}

    # RequestsOffCalendar (preferred)
    requests_off: dict[str, list[date]] = {}
    if "RequestsOffCalendar" in wb.sheetnames:
        ws = wb["RequestsOffCalendar"]
        # header: staff_id, name, 1..N
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if len(header) < 3 or str(header[0]).strip() != "staff_id":
            raise ValueError("RequestsOffCalendar のヘッダが不正です。")
        day_cols: list[int] = []
        for h in header[2:]:
            if h in (None, ""):
                continue
            day_cols.append(int(str(h).strip()))

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            sid = str(row[0]).strip()
            if sid not in staff_ids:
                raise ValueError(f"RequestsOffCalendar に未知の staff_id があります: {sid}")
            # cells from col 3 map to day numbers
            for i, day_num in enumerate(day_cols):
                cell = row[2 + i] if (2 + i) < len(row) else None
                if cell is None or cell == "":
                    continue
                sval = str(cell).strip().lower()
                if sval in ("off", "1", "x", "yes", "y", "true"):
                    d = _parse_date(f"{month}-{day_num:02d}")
                    requests_off.setdefault(sid, []).append(d)

    # RequestsOff (legacy, append)
    if "RequestsOff" in wb.sheetnames:
        ws = wb["RequestsOff"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            sid = str(row[0]).strip()
            if sid not in staff_ids:
                raise ValueError(f"RequestsOff に未知の staff_id があります: {sid}")
            d = _as_date_cell(row[1] if len(row) > 1 else None)
            if d is None:
                continue
            requests_off.setdefault(sid, []).append(d)

    # Closed
    closed_dates: list[date] = []
    if "Closed" in wb.sheetnames:
        ws = wb["Closed"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ""):
                continue
            d = _as_date_cell(row[0])
            if d is not None:
                closed_dates.append(d)

    return MonthInput(
        month=month,
        staff=tuple(staff),
        closed_dates=tuple(sorted(set(closed_dates))),
        requests_off={sid: tuple(sorted(set(ds))) for sid, ds in requests_off.items()},
        requirements=requirements,
        auto_close_jp_holidays=auto_close,
    )
